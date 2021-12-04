import xlrd
import pandas as pd
from openpyxl import load_workbook
from xlrd import open_workbook
import nltk
from nltk.tree import Tree
from nltk.parse.generate import generate
from nltk.tree import *
import os
from nltk.tokenize import word_tokenize
from nltk.tokenize import sent_tokenize
import xml.etree.ElementTree as etree
import xlrd
import time
import sys
from nltk import induce_pcfg
from nltk.parse import pchart
from nltk import PCFG
from nltk.draw.util import CanvasFrame


sys.setrecursionlimit(5000)

##start = time.time()
##PERIOD_OF_TIME = 15 # 5min
##while True :


sen = input("Enter your sentence: ")
sent = word_tokenize(sen)



#sen = "مهربانی وکړه بیاي ووايه . يوسف غلے شو . دیړ وخت وشو نہ خکاری"
##for i in sent_tokenize(sen):
##    print(i)

##S -> NP VP | NP PP PP VP | PP NP VP
##NP -> ADV N | N 
##VP -> NP VP | ADV VP | V
##ADV -> ADJ ADJ | AV NP

gram =("""
S -> NP VP 
NP -> ADJ | N | N N | PN | ADJ N | AV N | N ADJ | NU NU | NU AP | ADJ AP | AV | ADJ AP | N PN | VP N | PN ADV | AV ADV | N VP | NU N | NU | V | AV AP | ADJ VP | N AP  | ADJ AP  | ADJ NP  | N NP                    
VP -> V AP | ADJ V | AP | NP | AV PN | V ADV | V | AV AP | N ADV | N | NU N | N V | ADJ AP | N AV | V ADJ| ADJ NP | N AP  | N NP | NP NP| AV VP| ADJ VP | N VP 
AP -> AV V | V NP | ADJ V | NP VP | AV NP | PN NP | N V | NU N | AV N | ADJ PN | V VP | N ADV | PN AV | ADJ VP | PN N| AV ADV     
ADV -> ADV ADJ | PN VP | N AP | AV AV | V AP | N V 

""")





###د هغه ناوړه ملګري وویل
##gram = ("""
##S -> NP VP [1.0]
##NP -> AV [0.5] | ADJ AP [0.5] 
##VP -> AP [1.0]
##AP -> PN NP [0.5] | N V [0.5]
##AV -> "د" [1.0]
##PN -> "هغه" [1.0]
##ADJ -> "ناوړه" [1.0]
##V -> "وویل" [1.0]
##N -> "ملګري" [1.0]
##
##""")



##یوه وفاداره میرمن جوړه شوه
##gram = ("""
##S -> NP VP
##NP -> NU | N N
##VP -> NP NP
##
##""")


#دویم تن وویل
##gram =("""
##S -> NP VP
##NP -> V
##VP -> N V
##""")


book = open_workbook("Pastho dictionary2.xlsx")
for sheet in book.sheets():
    for rowidx in range(sheet.nrows):
        row = sheet.row(rowidx)
        for i in sent:
            for colidx,cell in enumerate(row):
                if cell.value == i:#row value                    
                    #print ("Found Row Element")
                    #print(rowidx, colidx)
                    #print(cell.value)
                    print(row)
                    print('\n')

book = load_workbook("Pastho dictionary2.xlsx")
worksheet = book.sheetnames
sheet = book["Sheet1"]
c=1


for i in sheet: 
    d = sheet.cell(row=c, column=2)

    if(d.value is None):
        print(" Try Again ")


    elif (d.value == "			Noun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "N ->" + "'" + cell.value + "'" + "\n"                    
            
        
    elif (d.value == "Noun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "N ->" + "'" + cell.value + "'" + "\n"    


    elif (d.value == "			Verb  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "V ->" + "'" + cell.value + "'" + "\n"


    elif (d.value == "Verb  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "V ->" + "'" + cell.value + "'" + "\n"
        

    elif (d.value == "			Adjective  "):
        
        cell = sheet.cell(row=c, column=1)
        gram = gram + "ADJ ->" + "'" + cell.value + "'" + "\n"


    elif (d.value == "Adjective  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "ADJ ->" + "'" + cell.value + "'" + "\n"
        

    elif (d.value == "			Participles  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "PP ->" + "'" + cell.value + "'" + "\n" 
        #print("hi")
    
    elif (d.value == "			Adverb  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "AV ->" + "'" + cell.value +  "'" + "\n"


    elif (d.value == "Adverb  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "AV ->" + "'" + cell.value +  "'" + "\n"   


    elif (d.value == "			numerical  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "NU ->" + "'" + cell.value + "'" + "\n"


    elif (d.value == "numerical  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "NU ->" + "'" + cell.value + "'" + "\n"


    elif (d.value == "			proNoun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "PN ->" + "'" + cell.value + "'" + "\n"



    elif (d.value == "			ProNoun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "PN ->" + "'" + cell.value + "'" + "\n"



    elif (d.value == "ProNoun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "PN ->" + "'" + cell.value + "'" + "\n"



    elif (d.value == "			suffix  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "SA ->" + "'" + cell.value + "'" + "\n"



    elif (d.value == "			Suffix  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "SA ->" + "'" + cell.value + "'" + "\n"

    c=c+1

#print(gram)
grammar1 = nltk.CFG.fromstring(gram)
sr_parser = nltk.parse.chart.ChartParser(grammar1)
#sr_parser = nltk.RecursiveDescentParser(grammar1)
#sr_parser = nltk.ShiftReduceParser(grammar1)

for tree in sr_parser.parse(sent):    
    print(tree)




##    values = tree
##    with open("new1.txt", "w", encoding='utf-8') as output:
##        output.write(str(values))
        






    #break
##    if time.time() > start + PERIOD_OF_TIME:
##        break

    
    

