import xlrd
import pandas as pd
from openpyxl import load_workbook
from xlrd import open_workbook
import nltk
from nltk.tree import Tree
from nltk.parse.generate import generate
from nltk.tree import *
import os
from nltk.tokenize import word_tokenize
from nltk.tokenize import sent_tokenize
import xml.etree.ElementTree as etree
import xlrd
import time
import sys
from nltk import induce_pcfg
from nltk.parse import pchart
from nltk import PCFG
from nltk.draw.util import CanvasFrame


sys.setrecursionlimit(5000)

##start = time.time()
##PERIOD_OF_TIME = 15 # 5min
##while True :


sen = input("Enter your sentence: ")
sent = word_tokenize(sen)



#sen = "مهربانی وکړه بیاي ووايه . يوسف غلے شو . دیړ وخت وشو نہ خکاری"
##for i in sent_tokenize(sen):
##    print(i)


gram =("""
S -> NP VP [1.0]
NP -> ADJ [0.0] | N [0.0] | N N [0.4] | PN [0.0]  | ADJ N [0.0] | AV N [0.0] | N ADJ [0.1] | NU NU [0.5] | NU AP [0.0] | ADJ AP [0.0] | AV [0.0] | ADJ AP [0.0] | N PN [0.0] | VP N [0.0] | PN ADV [0.0] | AV ADV [0.0] | N VP [0.0] | NU N [0.0] | NU [0.0] | V [0.0] | AV AP [0.0] | ADJ VP [0.0] | N AP [0.0] | ADJ AP [0.0] | ADJ NP [0.0] | N NP [0.0]                   
VP -> V AP [0.557] | ADJ V [0.05] | AP [0.0] | NP [0.0] | AV PN [0.056] | V ADV [0.0] | V [0.0] | AV AP [0.0] | N ADV [0.0] | N [0.0] | NU N [0.2] | N V [0.0375] | ADJ AP [0.0] | N AV [0.10] | V ADJ [0.0] | ADJ NP [0.0] | N AP [0.0] | N NP [0.0] | NP NP [0.0] | AV VP [0.0] | ADJ VP [0.0] | N VP [0.0]
AP -> AV V [0.056] | V NP [0.166] | ADJ V [0.051] | NP VP [0.0] | AV NP [0.0] | PN NP [0.0] | N V [0.037] | NU N [0.2] | AV N [0.2] | ADJ PN [0.066] | V VP [0.0] | N ADV [0.0] | PN AV [0.024] | ADJ VP [0.0] | PN N [0.2] | AV ADV [0.0]    
ADV -> ADV ADJ [0.5] | PN VP [0.0] | N AP [0.0] | AV AV [0.5] | V AP [0.0] | N V [0.0]
""")


###د هغه ناوړه ملګري وویل
##gram = ("""
##S -> NP VP [1.0]
##NP -> AV [0.5] | ADJ AP [0.5] 
##VP -> AP [1.0]
##AP -> PN NP [0.5] | N V [0.5]
##AV -> "د" [1.0]
##PN -> "هغه" [1.0]
##ADJ -> "ناوړه" [1.0]
##V -> "وویل" [1.0]
##N -> "ملګري" [1.0]
##
##""")



##یوه وفاداره میرمن جوړه شوه
##gram = ("""
##S -> NP VP
##NP -> NU | N N
##VP -> NP NP
##
##""")


#دویم تن وویل
##gram =("""
##S -> NP VP
##NP -> V
##VP -> N V
##""")


book = open_workbook("Pastho dictionary2.xlsx")
for sheet in book.sheets():
    for rowidx in range(sheet.nrows):
        row = sheet.row(rowidx)
        for i in sent:
            for colidx,cell in enumerate(row):
                if cell.value == i:#row value                    
                    #print ("Found Row Element")
                    #print(rowidx, colidx)
                    #print(cell.value)
                    print(row)
                    print('\n')

book = load_workbook("Pastho dictionary2.xlsx")
worksheet = book.sheetnames
sheet = book["Sheet1"]
c=1


for i in sheet: 
    d = sheet.cell(row=c, column=2)

    if(d.value is None):
        print(" Try Again ")


    elif (d.value == "			Noun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "N ->" + "'" + cell.value + "'" + " " + "[0.0000851934]" + "\n"                    
            
        
    elif (d.value == "Noun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "N ->" + "'" + cell.value + "'" + " " + "[0.0000851934]" + "\n"    


    elif (d.value == "			Verb  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "V ->" + "'" + cell.value + "'" + " " + "[0.0005530973]" + "\n"


    elif (d.value == "Verb  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "V ->" + "'" + cell.value + "'" + " " + "[0.0005530973]" + "\n"
        

    elif (d.value == "			Adjective  "):
        
        cell = sheet.cell(row=c, column=1)
        gram = gram + "ADJ ->" + "'" + cell.value + "'" + " " + "[0.000280112]" + "\n"


    elif (d.value == "Adjective  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "ADJ ->" + "'" + cell.value + "'" + " " + "[0.000280112]" + "\n"
        

    elif (d.value == "			Participles  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "PP ->" + "'" + cell.value + "'" + " " + "[0.0588235294]" + "\n" 
        #print("hi")
    
    elif (d.value == "			Adverb  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "AV ->" + "'" + cell.value +  "'" + " " + "[0.0025380711]" + "\n"


    elif (d.value == "Adverb  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "AV ->" + "'" + cell.value +  "'" + " " + "[0.0025380711]" + "\n"   


    elif (d.value == "			numerical  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "NU ->" + "'" + cell.value + "'" + " " + "[0.0222222222]" + "\n"


    elif (d.value == "numerical  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "NU ->" + "'" + cell.value + "'" + " " + "[0.0222222222]" + "\n"


    elif (d.value == "			proNoun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "PN ->" + "'" + cell.value + "'" + " " + "[0.0125]" + "\n"



    elif (d.value == "			ProNoun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "PN ->" + "'" + cell.value + "'" + " " + "[0.0125]" + "\n"



    elif (d.value == "ProNoun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "PN ->" + "'" + cell.value + "'" + " " + "[0.0125]" + "\n"



    elif (d.value == "			suffix  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "SA ->" + "'" + cell.value + "'" + " " + "[0.0476190476]" + "\n"



    elif (d.value == "			Suffix  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "SA ->" + "'" + cell.value + "'" + " " + "[0.0476190476]" + "\n"

    c=c+1

#print(gram)
grammar1 = nltk.PCFG.fromstring(gram)
sr_parser = nltk.RecursiveDescentParser(grammar1)
#sr_parser = nltk.ShiftReduceParser(grammar1)

for tree in sr_parser.parse(sent):    
    print(tree)




##    values = tree
##    with open("new1.txt", "w", encoding='utf-8') as output:
##        output.write(str(values))
        






    #break
##    if time.time() > start + PERIOD_OF_TIME:
##        break

    
    

