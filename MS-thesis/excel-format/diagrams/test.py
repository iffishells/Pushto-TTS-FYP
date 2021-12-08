import xlrd
import pandas as pd
from openpyxl import load_workbook
from xlrd import open_workbook
import nltk
from nltk.tree import Tree
from nltk.parse.generate import generate
from nltk.tree import *
import os
from nltk.tokenize import word_tokenize
from nltk.tokenize import sent_tokenize
import xml.etree.ElementTree as etree
import xlrd
import time
import sys
from nltk import induce_pcfg
from nltk.parse import pchart
from nltk import PCFG
from nltk.draw.util import CanvasFrame


sys.setrecursionlimit(5000)

##start = time.time()
##PERIOD_OF_TIME = 15 # 5min
##while True :


sen = input("Enter your sentence: ")
sent = word_tokenize(sen)



#sen = "مهربانی وکړه بیاي ووايه . يوسف غلے شو . دیړ وخت وشو نہ خکاری"
##for i in sent_tokenize(sen):
##    print(i)


gram =("""
S -> NP VP [1.0]
NP -> ADJ [0.1] | N [0.1] | N N [0.2] | PN [0.02] | ADJ N [0.1] | AV N [0.02] | N ADJ [0.1] | NU NU [0.1] | NU AP [0.02] | ADJ AP [0.02] | AV [0.02] | ADJ AP [0.02] | VP N [0.02] | AV ADV [0.01] | N VP [0.01] | NU N [0.02] | NU [0.01] | V [0.01] | AV AP [0.01] | ADJ VP [0.01] | N AP [0.01] | ADJ AP [0.01] | ADJ NP [0.01] | N NP [0.05]                  
VP -> V AP [0.1] | AP [0.1] | NP [0.1] | V ADV [0.1] | V [0.1] | AV AP [0.1] | N ADV [0.1] | NU N [0.1] | N V [0.01] | ADJ AP [0.01] | N AV [0.1] | V ADJ [0.01] | ADJ NP [0.01] | N AP [0.01] | N NP [0.01] | NP NP [0.01] | AV VP [0.01] | ADJ VP [0.01] | N VP [0.01]
AP -> AV V [0.02] | V NP [0.1] | ADJ V [0.1] | NP VP [0.1] | AV NP [0.02] | PN NP [0.02] | N V [0.08] | NU N [0.1] | AV N [0.1] | V VP [0.02] | N ADV [0.1] | PN AV [0.02] | ADJ VP [0.1] | PN N [0.1] | AV ADV [0.02]    
ADV -> ADV ADJ [0.2] | PN VP [0.2] | N AP [0.2] | AV AV [0.2] | V AP [0.1] | N V [0.1]
""")

##VP -> ADJ V [] 2 |  AV PN [0.0454545455]  4  | N [0.01]  8 |
##AP ->  1 |ADJ PN [0.0625]  9 | 

##
##د هغه ناوړه ملګري وویل
##
##gram = ("""
##S -> NP VP [1.0]
##NP -> AV [0.5] | ADJ AP [0.5] 
##VP -> AP [1.0]
##AP -> PN NP [0.5] | N V [0.5]
##AV -> "د" [1.0]
##PN -> "هغه" [1.0]
##ADJ -> "ناوړه" [1.0]
##V -> "وویل" [1.0]
##N -> "ملګري" [1.0]
##""")



##یوه وفاداره میرمن جوړه شوه
##gram = ("""
##S -> NP VP
##NP -> NU | N N
##VP -> NP NP
##
##""")


#دویم تن وویل
##gram =("""
##S -> NP VP
##NP -> V
##VP -> N V
##""")



##gram =("""
##S -> NP VP [1]
##NP -> ADJ [1]
##VP -> V [1]
##AP -> AV V [1]
##""")



book = open_workbook("Pastho dictionary2.xlsx")
for sheet in book.sheets():
    for rowidx in range(sheet.nrows):
        row = sheet.row(rowidx)
        for i in sent:
            for colidx,cell in enumerate(row):
                if cell.value == i:#row value                    
                    #print ("Found Row Element")
                    #print(rowidx, colidx)
                    #print(cell.value)
                    print(row)
                    print('\n')

book = load_workbook("Pastho dictionary2.xlsx")
worksheet = book.sheetnames
sheet = book["Sheet1"]
c=1


for i in sheet: 
    d = sheet.cell(row=c, column=2)

    if(d.value is None):
        print(" Try Again ")


    elif (d.value == "			Noun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "N ->" + "'" + cell.value + "'" + " " + "[0.0000851934]" + "\n"                    
            
        
    elif (d.value == "Noun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "N ->" + "'" + cell.value + "'" + " " + "[0.0000851934]" + "\n"    
#0.0005530973

    elif (d.value == "			Verb  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "V ->" + "'" + cell.value + "'" + " " + "[0.0005530973]" + "\n"


    elif (d.value == "Verb  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "V ->" + "'" + cell.value + "'" + " " + "[0.0005530973]" + "\n"
        

    elif (d.value == "			Adjective  "):
        
        cell = sheet.cell(row=c, column=1)
        gram = gram + "ADJ ->" + "'" + cell.value + "'" + " " + "[0.000280112]" + "\n"


    elif (d.value == "Adjective  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "ADJ ->" + "'" + cell.value + "'" + " " + "[0.000280112]" + "\n"
        

    elif (d.value == "			Participles  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "PP ->" + "'" + cell.value + "'" + " " + "[0.0588235294]" + "\n" 
        #print("hi")
    
    elif (d.value == "			Adverb  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "AV ->" + "'" + cell.value +  "'" + " " + "[0.0025380711]" + "\n"


    elif (d.value == "Adverb  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "AV ->" + "'" + cell.value +  "'" + " " + "[0.0025380711]" + "\n"   


    elif (d.value == "			numerical  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "NU ->" + "'" + cell.value + "'" + " " + "[0.0222222222]" + "\n"


    elif (d.value == "numerical  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "NU ->" + "'" + cell.value + "'" + " " + "[0.0222222222]" + "\n"


    elif (d.value == "			proNoun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "PN ->" + "'" + cell.value + "'" + " " + "[0.0125]" + "\n"



    elif (d.value == "			ProNoun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "PN ->" + "'" + cell.value + "'" + " " + "[0.0125]" + "\n"



    elif (d.value == "ProNoun  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "PN ->" + "'" + cell.value + "'" + " " + "[0.0125]" + "\n"



    elif (d.value == "			suffix  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "SA ->" + "'" + cell.value + "'" + " " + "[0.0476190476]" + "\n"



    elif (d.value == "			Suffix  "):
        cell = sheet.cell(row=c, column=1)
        gram = gram + "SA ->" + "'" + cell.value + "'" + " " + "[0.0476190476]" + "\n"

    c=c+1

#print(gram)
grammar1 = nltk.PCFG.fromstring(gram)
sr_parser = nltk.ViterbiParser(grammar1)

#max=0
for tree in sr_parser.parse(sent):
    print(tree)
    with open("prob.txt", "a", encoding='utf-8') as output:
        output.write(str(tree))
        output.write("\n")









##    if (tree.prob() > max):
##        max=tree.prob()
##        max_tree=tree
##
##print(max)
##print(max_tree)



##sr_parser = nltk.parse.chart.ChartParser(grammar1)
#sr_parser = nltk.RecursiveDescentParser(grammar1)
#sr_parser = nltk.ShiftReduceParser(grammar1)


##for tree in sr_parser.parse(sent):    
##    #values = tree
##    
##    with open("test.txt", "a", encoding='utf-8') as output:
##        output.write(str(tree))
##        output.write("\n")
##
##    print(tree)
##    #break
##    
        






    #break
##    if time.time() > start + PERIOD_OF_TIME:
##        break

    
    

